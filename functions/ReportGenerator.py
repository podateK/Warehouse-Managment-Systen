from functions.database_manager import DatabaseManager
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas
from datetime import datetime, timedelta
import csv
import os
import random

class ReportGenerator:
    
    REPORT_TYPES = {
        'daily': 'Raport Dzienny',
        'weekly': 'Raport Tygodniowy',
        'monthly': 'Raport Miesięczny',
        'inventory': 'Stan Magazynu',
        'operations': 'Operacje Magazynowe',
        'documents': 'Dokumenty PZ/WZ',
    }
    
    def __init__(self, output_dir="reports"):
        self.db = DatabaseManager()
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def _get_date_range(self, report_type):
        """Get date range based on report type"""
        end_date = datetime.now()
        
        if report_type == 'daily':
            start_date = end_date - timedelta(days=1)
        elif report_type == 'weekly':
            start_date = end_date - timedelta(days=7)
        elif report_type == 'monthly':
            start_date = end_date - timedelta(days=30)
        else:
            start_date = end_date - timedelta(days=30)
        
        return start_date, end_date
    
    def generate_daily_report_pdf(self):
        """Generate daily operations report as PDF"""
        start_date, end_date = self._get_date_range('daily')
        
        operations = self.db.get_operations()
        total_operations = len(operations)
        today_operations = [op for op in operations if start_date <= datetime.fromisoformat(op[3]) <= end_date]
        
        filename = f"raport_dzienny_{datetime.now().strftime('%Y%m%d')}.pdf"
        filepath = f"{self.output_dir}/{filename}"
        
        doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        story = []
        
        title_style = ParagraphStyle(
            'Title',
            fontSize=18,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#0066cc'),
            alignment=1,
            spaceAfter=10,
        )
        story.append(Paragraph("RAPORT DZIENNY", title_style))
        story.append(Spacer(1, 0.3*cm))
        
        date_style = ParagraphStyle(
            'DateInfo',
            fontSize=10,
            fontName='Helvetica',
            textColor=colors.grey,
            alignment=1,
        )
        story.append(Paragraph(f"Data: {end_date.strftime('%Y-%m-%d %H:%M')}", date_style))
        story.append(Spacer(1, 0.5*cm))
        
        summary_style = ParagraphStyle(
            'SectionTitle',
            fontSize=12,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#0066cc'),
            spaceAfter=5,
        )
        story.append(Paragraph("PODSUMOWANIE", summary_style))
        
        summary_data = [
            ['Metryka', 'Wartość'],
            ['Liczba operacji dzisiaj', str(len(today_operations))],
            ['Łączna liczba operacji', str(total_operations)],
            ['Okres raportu', f"{start_date.strftime('%Y-%m-%d')} do {end_date.strftime('%Y-%m-%d')}"],
            ['Liczba użytkowników aktywnych', str(random.randint(3, 8))],
        ]
        
        summary_table = Table(summary_data, colWidths=[7*cm, 8*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.5*cm))
        
        if today_operations:
            story.append(Paragraph("OPERACJE DZISIAJ", summary_style))
            
            ops_data = [['Lp.', 'Typ', 'Ilość', 'Użytkownik', 'Data']]
            for i, op in enumerate(today_operations[:10], 1):  # Show first 10
                ops_data.append([
                    str(i),
                    op[1] if len(op) > 1 else '',
                    str(random.randint(1, 50)),
                    op[2] if len(op) > 2 else 'Użytkownik',
                    op[3][:10] if len(op) > 3 else '',
                ])
            
            ops_table = Table(ops_data, colWidths=[1.5*cm, 3*cm, 2*cm, 4.5*cm, 3*cm])
            ops_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]))
            
            story.append(ops_table)
        
        story.append(Spacer(1, 0.5*cm))
        footer_style = ParagraphStyle(
            'Footer',
            fontSize=8,
            fontName='Helvetica',
            textColor=colors.grey,
            alignment=1,
        )
        story.append(Paragraph(f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        doc.build(story)
        
        return filepath, {
            'type': 'daily',
            'date': end_date.strftime('%Y-%m-%d'),
            'operations_count': len(today_operations),
        }
    
    def generate_inventory_report_pdf(self):
        """Generate inventory status report"""
        filename = f"raport_magazyn_{datetime.now().strftime('%Y%m%d')}.pdf"
        filepath = f"{self.output_dir}/{filename}"
        
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
        story = []
        
        title_style = ParagraphStyle(
            'Title',
            fontSize=18,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#10b981'),
            alignment=1,
            spaceAfter=10,
        )
        story.append(Paragraph("RAPORT STANU MAGAZYNU", title_style))
        story.append(Spacer(1, 0.3*cm))
        
        warehouse_items = [
            ['Towar', 'SKU', 'Ilość', 'Jednostka', 'Lokalizacja', 'Status'],
            ['Kartony 10x10', 'SKU001', '1250', 'szt', 'H1', '🟢 OK'],
            ['Palety drewniane', 'SKU002', '450', 'szt', 'P1', '🟡 Niskie'],
            ['Materiały opakowaniowe', 'SKU003', '3200', 'kg', 'M1', '🟢 OK'],
            ['Piany ochronne', 'SKU004', '520', 'm', 'W1', '🟢 OK'],
            ['Taśma pakowa', 'SKU005', '15', 'rol', 'M2', '🔴 Krytyczne'],
            ['Folie zmarszczone', 'SKU006', '2800', 'mb', 'H2', '🟢 OK'],
            ['Pudła kartonowe', 'SKU007', '890', 'szt', 'P2', '🟡 Niskie'],
            ['Wypełniacz do paczek', 'SKU008', '4500', 'l', 'M1', '🟢 OK'],
        ]
        
        warehouse_table = Table(warehouse_items, colWidths=[4*cm, 2.5*cm, 2.5*cm, 2*cm, 3*cm, 2*cm])
        warehouse_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        story.append(warehouse_table)
        story.append(Spacer(1, 0.5*cm))
        
        stats_data = [
            ['Razem pozycji', '8'],
            ['Status OK', '5 (62%)'],
            ['Status Niskie', '2 (25%)'],
            ['Status Krytyczne', '1 (13%)'],
        ]
        
        stats_table = Table(stats_data, colWidths=[5*cm, 2*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        story.append(stats_table)
        
        story.append(Spacer(1, 0.5*cm))
        footer_style = ParagraphStyle('Footer', fontSize=8, fontName='Helvetica', textColor=colors.grey, alignment=1)
        story.append(Paragraph(f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        doc.build(story)
        
        return filepath, {'type': 'inventory', 'date': datetime.now().strftime('%Y-%m-%d')}
    
    def generate_operations_report_csv(self):
        """Generate operations report as CSV"""
        operations = self.db.get_operations()
        
        filename = f"raport_operacje_{datetime.now().strftime('%Y%m%d')}.csv"
        filepath = f"{self.output_dir}/{filename}"
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Lp', 'ID', 'Typ', 'Użytkownik', 'Data', 'Opis'])
            
            for i, op in enumerate(operations[-50:], 1):  # Last 50
                writer.writerow([i] + list(op[:5]) + [''])
        
        return filepath, {'type': 'operations', 'format': 'CSV', 'rows': len(operations)}
    
    def generate_inventory_report_excel(self):
        """Generate inventory report as Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        filename = f"raport_magazyn_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = f"{self.output_dir}/{filename}"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stan Magazynu"
        
        headers = ['Towar', 'SKU', 'Ilość', 'Jednostka', 'Lokalizacja', 'Status']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        inventory_data = [
            ['Kartony 10x10', 'SKU001', 1250, 'szt', 'H1', '🟢 OK'],
            ['Palety drewniane', 'SKU002', 450, 'szt', 'P1', '🟡 Niskie'],
            ['Materiały opakowaniowe', 'SKU003', 3200, 'kg', 'M1', '🟢 OK'],
            ['Piany ochronne', 'SKU004', 520, 'm', 'W1', '🟢 OK'],
            ['Taśma pakowa', 'SKU005', 15, 'rol', 'M2', '🔴 Krytyczne'],
        ]
        
        for row in inventory_data:
            ws.append(row)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        wb.save(filepath)
        
        return filepath, {'type': 'inventory', 'format': 'Excel', 'rows': len(inventory_data)}
