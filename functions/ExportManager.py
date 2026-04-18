import csv
from datetime import datetime
import os

class ExportManager:
    
    def __init__(self, output_dir="exports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def export_to_csv(self, data, filename, headers=None):
        filepath = os.path.join(self.output_dir, f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            if headers:
                writer.writerow(headers)
            
            writer.writerows(data)
        
        return filepath
    
    def export_to_excel(self, data, filename, headers=None, sheet_name="Data", title=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        filepath = os.path.join(self.output_dir, f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if title:
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            start_row = 2
        else:
            start_row = 1
        
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            data_start_row = start_row + 1
        else:
            data_start_row = start_row
        
        for row_idx, row_data in enumerate(data, data_start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
                
                border = Border(
                    left=Side(style='thin', color='cccccc'),
                    right=Side(style='thin', color='cccccc'),
                    top=Side(style='thin', color='cccccc'),
                    bottom=Side(style='thin', color='cccccc')
                )
                cell.border = border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.properties.created = datetime.now()
        wb.properties.modified = datetime.now()
        
        wb.save(filepath)
        
        return filepath
    
    def export_documents(self, documents, export_format='csv'):
        """
        Export warehouse documents
        
        Args:
            documents (list): List of document dicts
            export_format (str): 'csv' or 'excel'
            
        Returns:
            str: Path to exported file
        """
        headers = ['Numer', 'Typ', 'Data', 'Status', 'Wartość (PLN)', 'Przedmiot']
        
        data = []
        for doc in documents:
            data.append((
                doc.get('number', ''),
                doc.get('type', ''),
                doc.get('date', ''),
                doc.get('status', ''),
                f"{float(doc.get('value', 0)):.2f}",
                doc.get('subject', ''),
            ))
        
        if export_format == 'excel':
            return self.export_to_excel(
                data=data,
                filename='dokumenty',
                headers=headers,
                sheet_name='Dokumenty',
                title='DOKUMENTY MAGAZYNOWE'
            )
        else:
            return self.export_to_csv(
                data=data,
                filename='dokumenty',
                headers=headers
            )
    
    def export_inventory(self, items, export_format='csv'):
        """
        Export inventory data
        
        Args:
            items (list): List of inventory items
            export_format (str): 'csv' or 'excel'
            
        Returns:
            str: Path to exported file
        """
        headers = ['SKU', 'Nazwa', 'Ilość', 'Jednostka', 'Pozycja', 'Status']
        
        data = []
        for item in items:
            data.append((
                item.get('sku', ''),
                item.get('name', ''),
                item.get('quantity', ''),
                item.get('unit', ''),
                item.get('location', ''),
                item.get('status', ''),
            ))
        
        if export_format == 'excel':
            return self.export_to_excel(
                data=data,
                filename='magazyn',
                headers=headers,
                sheet_name='Magazyn',
                title='STAN MAGAZYNU'
            )
        else:
            return self.export_to_csv(
                data=data,
                filename='magazyn',
                headers=headers
            )
